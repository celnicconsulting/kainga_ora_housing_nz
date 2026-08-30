# ====================IMPORTS====================
"""Kāinga Ora / NZ Public Housing Intelligence.

This file is intended to read as a VISUAL SPECIFICATION of the application: the
comments describe what appears on screen, in what order, and what each colour and
marker means, so the interface can be reviewed and rebuilt from the source alone
without running it. Every `render_*` docstring is a layout description, top to
bottom, of the block it draws.

====================================================================
SCREEN LAYOUT
====================================================================

    ┌──────────────┬───────────────────────────────────────────────┐
    │  SIDEBAR     │  H1  Kāinga Ora / New Zealand Public Housing  │
    │              │      Intelligence                             │
    │ 🏘️ title      │                                               │
    │ caption      │  ╔═══════════════════════════════════════════╗ │
    │              │  ║ HAZARD BANNER (yellow/black, above tabs)  ║ │
    │ Latest       │  ║ NOT AN OFFICIAL GOVERNMENT PRODUCT        ║ │
    │ published    │  ╚═══════════════════════════════════════════╝ │
    │ periods      │                                               │
    │  · KO stock  │  ┌ tab bar ──────────────────────────────────┐ │
    │  · Register  │  │ 📊 │ 🗺️ │ 💰 │ 📋 │ 🔧 │ ⚙️              │ │
    │  · Rent      │  └───────────────────────────────────────────┘ │
    │  · HUD       │                                               │
    │              │  active tab body                              │
    │ Real vs      │                                               │
    │ synthetic    │                                               │
    │              │                                               │
    │ ℹ️ about      │                                               │
    └──────────────┴───────────────────────────────────────────────┘

Tabs, left to right:

    1  📊 National Overview   real       supply, demand, market context
    2  🗺️ Stock Map           part syn   H3 hexagons, resolutions 8-12
    3  💰 Market Rent vs IRR  mixed      the subsidy gap - the centrepiece
    4  📋 Housing Register    part syn   waitlist and the bedroom mismatch
    5  🔧 Asset & Maintenance full syn   condition, cost, backlog
    6  ⚙️ Pipeline            real       provenance, lineage, reconciliation

====================================================================
VISUAL VOCABULARY - the same meaning everywhere
====================================================================

    🔶  orange diamond      this figure is synthetic or synthetic-derived
    yellow/black stripes    provenance warning; read before any figure
    st.warning (amber)      partly synthetic tab - some series are real
    st.error (red)          fully synthetic tab - nothing here is measured
    BRAND blue   #2E86AB    measured quantities: stock, market rent, supply
    ACCENT red   #E4572E    pressure and demand: register, subsidy, mould
    INK navy     #1A3A5C    reference marks and tooltips
    blue gradient           density - more is simply more
    red-to-yellow           pressure - red is worse
    📋 heading + 📥 Excel   every detail table, button right-justified

Charts are Plotly; maps are pydeck H3HexagonLayer on a CartoDB Voyager basemap.
All money is NZD per week, the New Zealand convention. NZ English throughout.

====================================================================
CODE STRUCTURE
====================================================================

Follows the `snowflake-streamlit-development` template - cached `get_*` data
methods, `render_*` visual methods, a thin `main()`. The one departure is the
session layer: this build runs on Streamlit Community Cloud rather than
Streamlit in Snowflake, so `get_active_session()` is replaced by a read-only
DuckDB connection behind the same function boundary. Every data method keeps its
`df_db_schema` argument, so moving the app into Snowflake means swapping
`run_query` for `session.sql(...).to_pandas()` and nothing else.
"""

import io
import os
from pathlib import Path

import duckdb
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import pydeck as pdk
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    layout="wide",
    page_title="Kāinga Ora | NZ Public Housing Intelligence",
    page_icon="🏘️",
)

# ====================SESSION====================
APP_DIR = Path(__file__).resolve().parent
DB_CANDIDATES = [
    APP_DIR.parent / "data" / "kainga_ora_housing_public.duckdb",
    APP_DIR.parent / "public" / "kainga_ora_housing_public.duckdb",
    Path("data/kainga_ora_housing_public.duckdb"),
]

# Kept for parity with the Snowflake template; unused against DuckDB, where the
# schemas are attached in the same file.
df_db = "PUBLIC"
df_schema_name = "MART"
df_db_schema = f"{df_db}.{df_schema_name}"

# Palette. The three colours carry meaning and are used consistently: a reader
# who learns them on one tab can read the next without a legend.
BRAND = "#2E86AB"    # measured supply - Kainga Ora stock, market rent
INK = "#1A3A5C"      # reference marks, tooltip backgrounds, register totals
ACCENT = "#E4572E"   # pressure and demand - register, subsidy gap, mould

# Map opens on Wellington CBD, as the work packet specifies. The quick-jump
# buttons fly the camera to the three main centres; zoom is tuned per centre so
# each frames its own urban area rather than sharing one national zoom.
WELLINGTON = (-41.2865, 174.7762)
QUICK_JUMPS = {
    "Wellington": (-41.2865, 174.7762, 11.0),
    "Auckland": (-36.8485, 174.7633, 10.0),
    "Christchurch": (-43.5321, 172.6362, 11.0),
    "All of New Zealand": (-41.0, 173.0, 4.8),
}


def _extract_fingerprint():
    """Size and modified time of the extract, used as a cache key.

    Streamlit Community Cloud hot-reloads on a push: it pulls the new files and
    re-runs the script, but does NOT clear `cache_resource`. A connection opened
    before the pull therefore keeps reading the replaced file, and a query
    written against a newly added column fails to bind against data that is
    sitting correctly on disk a few bytes away.

    Making the fingerprint an argument to `get_connection` turns a data refresh
    into a cache miss, so the connection reopens by itself. Without it the app
    needs a manual reboot after every data change, which is a step that will
    eventually be forgotten.
    """
    for path in DB_CANDIDATES:
        if path.exists():
            stat = path.stat()
            return (str(path), stat.st_size, int(stat.st_mtime))
    return ("missing", 0, 0)


@st.cache_resource
def _open_connection(fingerprint):
    """Open the extract read-only. Re-runs whenever the fingerprint changes."""
    path = fingerprint[0]
    if path == "missing":
        raise FileNotFoundError(
            "kainga_ora_housing_public.duckdb not found. Expected it beside the "
            "app under data/. Run scripts/09_build_public.py to create it."
        )
    return duckdb.connect(path, read_only=True)


def get_connection():
    """The read-only connection to the published extract.

    Re-opens automatically when the extract on disk changes; see
    `_extract_fingerprint` for why that matters on Community Cloud.
    """
    return _open_connection(_extract_fingerprint())


@st.cache_data(show_spinner=False)
def _run_query_cached(sql: str, fingerprint) -> pd.DataFrame:
    """Execute and cache, keyed on both the SQL and the extract it ran against.

    `fingerprint` is unused in the body and essential in the signature: it is
    part of the cache key, so refreshing the data file invalidates every cached
    frame rather than leaving rows from the previous extract in memory.
    """
    return _open_connection(fingerprint).execute(sql).df()


def run_query(sql: str) -> pd.DataFrame:
    """Run SQL against the extract and cache the frame.

    The single seam between the app and its data. Swapping this one function for
    `session.sql(sql).to_pandas()` moves the whole app into Streamlit in
    Snowflake; nothing else in the file knows which engine it is talking to.
    """
    return _run_query_cached(sql, _extract_fingerprint())


# ====================DATA====================
# One cached query per visual element. Each docstring names the element it
# feeds, so any chart on screen can be traced back to the SQL behind it and any
# query here can be located on screen. Nothing is computed twice: aggregation
# happens in DuckDB, the app only draws.
@st.cache_data(show_spinner=False)
def get_national_quarter(df_db_schema):
    """Feeds: Tab 1 - supply and register trend lines, and the detail table."""
    return run_query("""
        SELECT * FROM MART.M_NATIONAL_QUARTER
        WHERE PERIOD >= '2014Q1'
        ORDER BY PERIOD
    """)


@st.cache_data(show_spinner=False)
def get_latest_snapshot(df_db_schema):
    """Feeds: Tab 1 - the five-metric KPI row. Sidebar - latest published
    periods."""
    return run_query("""
        WITH ko AS (
            SELECT PERIOD, STATE_RENTALS FROM MART.M_KO_STOCK_NATIONAL_QUARTER
            ORDER BY PERIOD DESC LIMIT 1
        ), reg AS (
            SELECT PERIOD, REGISTER_TOTAL, PRIORITY_A, PRIORITY_B, TRANSFER_TOTAL
            FROM MART.M_REGISTER_NATIONAL_QUARTER
            WHERE REGISTER_TOTAL IS NOT NULL ORDER BY PERIOD DESC LIMIT 1
        ), chp AS (
            SELECT PERIOD, VALUE AS CHP_HOMES FROM MART.M_HUD_HEADLINE
            WHERE METRIC = 'CHP_HOMES' ORDER BY PERIOD DESC LIMIT 1
        ), rent AS (
            SELECT PERIOD, MEDIAN_RENT, GEOMETRIC_MEAN_RENT
            FROM MART.M_BOND_NATIONAL_QUARTER
            WHERE MEDIAN_RENT IS NOT NULL ORDER BY PERIOD DESC LIMIT 1
        )
        SELECT
            ko.PERIOD          AS KO_PERIOD,
            ko.STATE_RENTALS,
            reg.PERIOD         AS REGISTER_PERIOD,
            reg.REGISTER_TOTAL, reg.PRIORITY_A, reg.PRIORITY_B, reg.TRANSFER_TOTAL,
            chp.PERIOD         AS CHP_PERIOD,
            chp.CHP_HOMES,
            rent.PERIOD        AS RENT_PERIOD,
            rent.MEDIAN_RENT, rent.GEOMETRIC_MEAN_RENT
        FROM ko, reg, chp, rent
    """)


@st.cache_data(show_spinner=False)
def get_subsidy_summary(df_db_schema):
    """Feeds: Tab 1 - the 🔶 weekly IRRS subsidy metric and its caption."""
    return run_query("""
        SELECT
            COUNT(*)                              AS TENANCIES,
            ROUND(AVG(MARKET_RENT_WEEKLY), 2)     AS AVG_MARKET_RENT,
            ROUND(AVG(INCOME_RELATED_RENT_WEEKLY), 2) AS AVG_IRR,
            ROUND(AVG(IRRS_SUBSIDY_WEEKLY), 2)    AS AVG_SUBSIDY,
            ROUND(SUM(IRRS_SUBSIDY_WEEKLY), 0)    AS TOTAL_WEEKLY_SUBSIDY,
            ROUND(SUM(IRRS_SUBSIDY_WEEKLY) * 52.0 / 1e9, 2) AS ANNUAL_SUBSIDY_BN
        FROM SYN.SYN_KO_TENANCY
    """)


@st.cache_data(show_spinner=False)
def get_vacancy_summary(df_db_schema):
    """Feeds: Tab 1 - vacancy share behind the KPI row."""
    return run_query("""
        SELECT
            SUM(CASE WHEN STATUS <> 'Tenanted' THEN 1 ELSE 0 END) AS VACANT,
            COUNT(*)                                              AS TOTAL,
            ROUND(100.0 * SUM(CASE WHEN STATUS <> 'Tenanted' THEN 1 ELSE 0 END)
                  / COUNT(*), 2)                                  AS VACANCY_PCT
        FROM SYN.SYN_KO_PROPERTY
    """)


@st.cache_data(show_spinner=False)
def get_stock_portfolio(df_db_schema):
    """Feeds: Tab 1 - stacked area, portfolio composition."""
    return run_query("""
        SELECT PERIOD, PORTFOLIO, SUM(PROPERTIES) AS PROPERTIES
        FROM MART.M_KO_STOCK_PORTFOLIO
        GROUP BY ALL
        ORDER BY PERIOD, PORTFOLIO
    """)


@st.cache_data(show_spinner=False)
def get_map_data(df_db_schema, resolution, ta_filter):
    """Feeds: Tab 2 - the H3 hexagon layer, its four metrics and detail table."""
    hex_col = f"H3_RES_{resolution}"
    where = ""
    if ta_filter and ta_filter != "All of New Zealand":
        safe = str(ta_filter).replace("'", "''")
        where = f"WHERE p.TA_NAME = '{safe}'"
    return run_query(f"""
        SELECT
            p.{hex_col}                              AS HEX,
            COUNT(*)                                 AS PROPERTY_COUNT,
            ROUND(AVG(p.CONDITION_SCORE), 2)         AS MEAN_CONDITION,
            ROUND(100.0 * SUM(CASE WHEN p.STATUS <> 'Tenanted' THEN 1 ELSE 0 END)
                  / COUNT(*), 2)                     AS VACANCY_RATE,
            ROUND(AVG(t.IRRS_SUBSIDY_WEEKLY), 2)     AS MEAN_SUBSIDY,
            ROUND(AVG(CAST(NULLIF(REPLACE(p.BEDROOMS, '5+', '5'), 'Bedsit')
                           AS DOUBLE)), 2)           AS MEAN_BEDROOMS,
            MODE(p.TA_NAME)                          AS DOMINANT_TA,
            MODE(p.DWELLING_TYPE)                    AS DOMINANT_DWELLING
        FROM SYN.SYN_KO_PROPERTY p
        LEFT JOIN SYN.SYN_KO_TENANCY t ON t.PROPERTY_ID = p.PROPERTY_ID
        {where}
        GROUP BY ALL
    """)


@st.cache_data(show_spinner=False)
def get_ta_options(df_db_schema):
    """Feeds: Tab 2 - the territorial authority filter."""
    df = run_query("""
        SELECT DISTINCT TA_NAME FROM SYN.SYN_KO_PROPERTY ORDER BY TA_NAME
    """)
    return ["All of New Zealand"] + df["TA_NAME"].tolist()


@st.cache_data(show_spinner=False)
def get_rent_comparison(df_db_schema, period, bedrooms):
    """Feeds: Tab 3 - overlaid bars, subsidy scatter and the detail table."""
    safe_period = str(period).replace("'", "''")
    safe_beds = str(bedrooms).replace("'", "''")
    return run_query(f"""
        WITH market AS (
            SELECT TA_NAME, REGION, TA_RENT_MEASURED, MARKET_RENT_DERIVED
            FROM MART.M_MARKET_RENT_TA_BEDROOM
            WHERE PERIOD = '{safe_period}' AND BEDROOMS = '{safe_beds}'
        ), bands AS (
            SELECT TA_NAME, LOWER_QUARTILE_RENT, UPPER_QUARTILE_RENT,
                   MEDIAN_RENT, ACTIVE_BONDS
            FROM MART.M_BOND_TA_QUARTER WHERE PERIOD = '{safe_period}'
        ), irr AS (
            SELECT TA_NAME,
                   COUNT(*)                                  AS TENANCIES,
                   ROUND(AVG(INCOME_RELATED_RENT_WEEKLY), 2) AS MEAN_IRR,
                   ROUND(AVG(IRRS_SUBSIDY_WEEKLY), 2)        AS MEAN_SUBSIDY
            FROM SYN.SYN_KO_TENANCY
            WHERE BEDROOMS = '{safe_beds}'
            GROUP BY ALL
        )
        SELECT m.TA_NAME, m.REGION,
               m.MARKET_RENT_DERIVED AS MARKET_RENT,
               m.TA_RENT_MEASURED,
               b.LOWER_QUARTILE_RENT, b.UPPER_QUARTILE_RENT, b.MEDIAN_RENT,
               b.ACTIVE_BONDS,
               i.TENANCIES, i.MEAN_IRR, i.MEAN_SUBSIDY
        FROM market m
        LEFT JOIN bands b USING (TA_NAME)
        LEFT JOIN irr i USING (TA_NAME)
        WHERE i.TENANCIES IS NOT NULL
        ORDER BY m.MARKET_RENT_DERIVED DESC
    """)


@st.cache_data(show_spinner=False)
def get_rent_periods(df_db_schema):
    """Feeds: Tab 3 - the quarter selector."""
    df = run_query("""
        SELECT DISTINCT PERIOD FROM MART.M_MARKET_RENT_TA_BEDROOM
        ORDER BY PERIOD DESC
    """)
    return df["PERIOD"].tolist()


@st.cache_data(show_spinner=False)
def get_rent_trend(df_db_schema, ta_names):
    """Feeds: Tab 3 - measured market rent over time, one line per district."""
    if not ta_names:
        return pd.DataFrame()
    quoted = ", ".join("'" + str(t).replace("'", "''") + "'" for t in ta_names)
    return run_query(f"""
        SELECT PERIOD, TA_NAME, GEOMETRIC_MEAN_RENT, MEDIAN_RENT,
               LOWER_QUARTILE_RENT, UPPER_QUARTILE_RENT
        FROM MART.M_BOND_TA_QUARTER
        WHERE TA_NAME IN ({quoted}) AND PERIOD >= '2000Q1'
        ORDER BY PERIOD
    """)


@st.cache_data(show_spinner=False)
def get_register_trend(df_db_schema):
    """Feeds: Tab 4 - the KPI row and the priority stacked area."""
    return run_query("""
        SELECT PERIOD, REGISTER_TOTAL, PRIORITY_A, PRIORITY_B, TRANSFER_TOTAL
        FROM MART.M_REGISTER_NATIONAL_QUARTER
        WHERE REGISTER_TOTAL IS NOT NULL
        ORDER BY PERIOD
    """)


@st.cache_data(show_spinner=False)
def get_register_by_region(df_db_schema):
    """Feeds: Tab 4 - register by region lines, and the region filter."""
    return run_query("""
        SELECT PERIOD, REGION, SUM(APPLICANTS) AS APPLICANTS
        FROM MART.M_REGISTER_TA_QUARTER
        WHERE REGION IS NOT NULL AND REGION <> '' AND APPLICANTS IS NOT NULL
        GROUP BY ALL
        ORDER BY PERIOD
    """)


@st.cache_data(show_spinner=False)
def get_bedroom_mismatch(df_db_schema):
    """Register demand by bedrooms required against Kāinga Ora bedroom supply."""
    return run_query("""
        WITH demand AS (
            SELECT BEDROOMS_REQUIRED AS BEDROOMS, SUM(APPLICANTS) AS APPLICANTS
            FROM MART.M_REGISTER_TA_BEDROOMS
            WHERE PERIOD = (SELECT MAX(PERIOD) FROM MART.M_REGISTER_TA_BEDROOMS)
              AND APPLICANTS IS NOT NULL
            GROUP BY ALL
        ), supply AS (
            SELECT CASE WHEN BEDROOMS = 'Bedsit' THEN '1' ELSE BEDROOMS END
                       AS BEDROOMS,
                   SUM(PROPERTIES) AS PROPERTIES
            FROM MART.M_KO_STOCK_TA_BEDROOM
            WHERE PERIOD = (SELECT MAX(PERIOD) FROM MART.M_KO_STOCK_TA_BEDROOM)
            GROUP BY ALL
        )
        SELECT COALESCE(d.BEDROOMS, s.BEDROOMS) AS BEDROOMS,
               d.APPLICANTS, s.PROPERTIES,
               ROUND(100.0 * d.APPLICANTS / NULLIF(
                   (SELECT SUM(APPLICANTS) FROM demand), 0), 1) AS DEMAND_PCT,
               ROUND(100.0 * s.PROPERTIES / NULLIF(
                   (SELECT SUM(PROPERTIES) FROM supply), 0), 1) AS SUPPLY_PCT
        FROM demand d FULL OUTER JOIN supply s USING (BEDROOMS)
        WHERE COALESCE(d.BEDROOMS, s.BEDROOMS) NOT IN ('Unknown')
        ORDER BY 1
    """)


@st.cache_data(show_spinner=False)
def get_register_ta_latest(df_db_schema):
    """Feeds: Tab 4 - register totals by district."""
    return run_query("""
        SELECT TA_NAME, REGION, APPLICANTS, LATITUDE, LONGITUDE
        FROM MART.M_REGISTER_TA_QUARTER
        WHERE PERIOD = (SELECT MAX(PERIOD) FROM MART.M_REGISTER_TA_QUARTER)
          AND APPLICANTS IS NOT NULL AND TA_NAME <> 'Unknown'
        ORDER BY APPLICANTS DESC
    """)


@st.cache_data(show_spinner=False)
def get_register_explorer(df_db_schema, region, priority):
    """Feeds: Tab 4 - the synthetic applicant detail table."""
    clauses = []
    if region and region != "All":
        clauses.append(f"REGION = '{str(region).replace(chr(39), chr(39) * 2)}'")
    if priority and priority != "All":
        clauses.append(f"PRIORITY = '{str(priority).replace(chr(39), chr(39) * 2)}'")
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    return run_query(f"""
        SELECT APPLICATION_ID, TA_NAME, REGION, PRIORITY, BEDROOMS_REQUIRED,
               HOUSEHOLD_TYPE, MAIN_APPLICANT_AGE_BAND, ETHNICITY, STATUS,
               LODGED_DATE, DAYS_ON_REGISTER, DAYS_TO_HOUSE
        FROM SYN.SYN_HOUSING_REGISTER
        {where}
        ORDER BY DAYS_ON_REGISTER DESC
        LIMIT 5000
    """)


@st.cache_data(show_spinner=False)
def get_days_to_house(df_db_schema):
    """Feeds: Tab 4 - days-to-house histogram and its median caption."""
    return run_query("""
        SELECT DAYS_TO_HOUSE, PRIORITY, REGION
        FROM SYN.SYN_HOUSING_REGISTER
        WHERE DAYS_TO_HOUSE IS NOT NULL
    """)


@st.cache_data(show_spinner=False)
def get_condition_by_decade(df_db_schema):
    """Feeds: Tab 5 - condition by build decade heatmap."""
    return run_query("""
        SELECT BUILD_DECADE, CONDITION_SCORE, COUNT(*) AS PROPERTIES
        FROM SYN.SYN_KO_PROPERTY
        GROUP BY ALL
        ORDER BY BUILD_DECADE, CONDITION_SCORE
    """)


@st.cache_data(show_spinner=False)
def get_maintenance_summary(df_db_schema):
    """Feeds: Tab 5 - the KPI row and cost-by-category bars."""
    return run_query("""
        SELECT CATEGORY, URGENCY,
               COUNT(*)                       AS WORK_ORDERS,
               ROUND(SUM(COST_NZD), 0)        AS TOTAL_COST,
               ROUND(AVG(COST_NZD), 2)        AS MEAN_COST,
               ROUND(AVG(DAYS_TO_CLOSE), 1)   AS MEAN_DAYS_TO_CLOSE
        FROM SYN.SYN_MAINTENANCE
        GROUP BY ALL
        ORDER BY TOTAL_COST DESC
    """)


@st.cache_data(show_spinner=False)
def get_maintenance_by_ta(df_db_schema):
    """Feeds: Tab 5 - the maintenance detail table."""
    return run_query("""
        SELECT m.TA_NAME, m.REGION,
               COUNT(*)                                  AS WORK_ORDERS,
               ROUND(SUM(m.COST_NZD), 0)                 AS TOTAL_COST,
               ROUND(SUM(m.COST_NZD) / COUNT(DISTINCT m.PROPERTY_ID), 2)
                                                         AS COST_PER_PROPERTY,
               ROUND(AVG(m.CONDITION_SCORE), 2)          AS MEAN_CONDITION,
               SUM(CASE WHEN m.CATEGORY = 'Mould-damp' THEN 1 ELSE 0 END)
                                                         AS MOULD_DAMP_ORDERS
        FROM SYN.SYN_MAINTENANCE m
        GROUP BY ALL
        ORDER BY TOTAL_COST DESC
    """)


@st.cache_data(show_spinner=False)
def get_mould_hotspots(df_db_schema):
    """Feeds: Tab 5 - the mould and damp hexagon map."""
    return run_query("""
        SELECT H3_RES_8 AS HEX,
               COUNT(*)                     AS MOULD_ORDERS,
               ROUND(SUM(COST_NZD), 0)      AS TOTAL_COST,
               MODE(TA_NAME)                AS DOMINANT_TA
        FROM SYN.SYN_MAINTENANCE
        WHERE CATEGORY = 'Mould-damp'
        GROUP BY ALL
        HAVING COUNT(*) >= 2
    """)


@st.cache_data(show_spinner=False)
def get_backlog_ageing(df_db_schema):
    """Feeds: Tab 5 - backlog ageing stacked bars."""
    return run_query("""
        SELECT CASE
                   WHEN DAYS_TO_CLOSE <= 7   THEN '0-7 days'
                   WHEN DAYS_TO_CLOSE <= 30  THEN '8-30 days'
                   WHEN DAYS_TO_CLOSE <= 90  THEN '31-90 days'
                   WHEN DAYS_TO_CLOSE <= 180 THEN '91-180 days'
                   ELSE '180+ days'
               END AS AGE_BAND,
               URGENCY,
               COUNT(*) AS WORK_ORDERS
        FROM SYN.SYN_MAINTENANCE
        GROUP BY ALL
    """)


@st.cache_data(show_spinner=False)
def get_source_register(df_db_schema):
    """Feeds: Tab 6 - the source register table."""
    return run_query("""
        SELECT DATASET_ID, SOURCE_NAME, KIND, CADENCE, LICENCE, TARGET_SCHEMA,
               URL, NOTES
        FROM META.SOURCE_REGISTER
        ORDER BY DATASET_ID, SOURCE_NAME
    """)


@st.cache_data(show_spinner=False)
def get_lineage(df_db_schema):
    """Feeds: Tab 6 - the real versus synthetic lineage table."""
    return run_query("SELECT * FROM META.LINEAGE ORDER BY BASIS, TABLE_NAME")


@st.cache_data(show_spinner=False)
def get_validation(df_db_schema):
    """Feeds: Tab 6 - the reconciliation checks table and its pass/fail state."""
    return run_query("""
        SELECT CHECK_NAME, STATUS, DETAIL, NOTE FROM STG.VALIDATION_RESULTS
        ORDER BY STATUS, CHECK_NAME
    """)


@st.cache_data(show_spinner=False)
def get_download_manifest(df_db_schema):
    """Feeds: Tab 6 - every file downloaded, with download date and MD5 checksum.
    This is the table the hazard banner points at."""
    return run_query("""
        SELECT dataset_id AS DATASET_ID, family AS FAMILY, period AS PERIOD,
               file_name AS FILE_NAME, ext AS FORMAT,
               ROUND(bytes / 1024.0, 1) AS SIZE_KB,
               downloaded_at AS DOWNLOADED_AT,
               md5 AS CHECKSUM_MD5,
               licence AS LICENCE,
               url AS SOURCE_URL
        FROM META.DOWNLOAD_MANIFEST
        ORDER BY dataset_id, period DESC, file_name
    """)


@st.cache_data(show_spinner=False)
def get_family_aliases(df_db_schema):
    """Feeds: Tab 6 - publication renames, every merge and deliberate non-merge."""
    return run_query("""
        SELECT OBSERVED_FAMILY, CANONICAL_FAMILY, ACTION, RATIONALE
        FROM META.FAMILY_ALIASES ORDER BY ACTION, OBSERVED_FAMILY
    """)


@st.cache_data(show_spinner=False)
def get_coverage(df_db_schema):
    """Feeds: Tab 6 - period coverage by publication family."""
    return run_query("""
        SELECT SCHEMA_NAME, FAMILY, CADENCE,
               COUNT(*)                                        AS PERIOD_SLOTS,
               SUM(CASE WHEN STATUS = 'PRESENT' THEN 1 ELSE 0 END) AS PRESENT,
               MIN(PERIOD) AS FIRST_PERIOD, MAX(PERIOD) AS LAST_PERIOD
        FROM META.DATA_GAPS
        GROUP BY ALL
        ORDER BY SCHEMA_NAME, FAMILY
    """)


# ====================REFERENCE_DOCUMENTS====================
# Long-form documentation is loaded from markdown files at run time rather than
# embedded as string literals in this module. Two reasons: the documents are
# reviewed and edited as documents, by people who should not have to touch Python
# to do it; and a 400-line string literal in the middle of an application file
# obscures the code around it. The files ship in the repository beside the app.

# The Build Notes file sits beside the application file and shares its stem, so
# `app/<name>.py` is documented by `app/<name>__readme.md`. Keeping them together
# means the pair is copied, reviewed and deployed as one unit; a document kept at
# the repository root drifts away from the code it describes.
BUILD_NOTES_PATH = APP_DIR / (Path(__file__).stem + "__readme.md")

REFERENCE_DOCS = {
    "build_notes": [
        BUILD_NOTES_PATH,
        APP_DIR / "kainga_ora_housing_nz__readme.md",
    ],
}


def _reference_doc_fingerprint(key: str):
    """Path, size and mtime of a reference document, used as a cache key.

    The same trap as the database connection: Community Cloud hot-reloads on a
    push but does not clear the caches, so a document cached before the pull
    would keep rendering while its file on disk had changed. Editing only the
    markdown - which is exactly what a wording change is - would then appear to
    do nothing. Recomputed on every run; never cached itself.
    """
    for path in REFERENCE_DOCS.get(key, []):
        try:
            if path.exists():
                stat = path.stat()
                return (str(path), stat.st_size, int(stat.st_mtime))
        except OSError:
            continue
    return ("missing", 0, 0)


@st.cache_data(show_spinner=False)
def _read_reference_doc(fingerprint) -> str:
    """Read the document. Re-runs whenever the file on disk changes."""
    path = fingerprint[0]
    if path == "missing":
        return ""
    try:
        return Path(path).read_text(encoding="utf-8")
    except OSError:
        return ""


def load_reference_doc(key: str) -> str:
    """Read a reference markdown document from disk.

    Returns the document text, or an empty string if no candidate path exists.
    Cached against the file's fingerprint, so editing the markdown is enough to
    change what the tab renders.

    Feeds: Tab 7 - the Build Notes document.
    """
    return _read_reference_doc(_reference_doc_fingerprint(key))


def render_tab_build_notes():
    """TAB 7 - Build Notes. The Phase Two write-up, rendered as markdown.

        H2  Build Notes                            + caption
        📥 Download the markdown                   right-justified button
        --- rule ---
        the document itself, rendered as markdown

    The content is NOT written in this file. It is the `__readme.md` beside it,
    loaded
    from disk at run time and rendered with `st.markdown`, so the document stays
    a document: reviewable in a pull request, readable on GitHub, and editable
    without touching the application.

    The heading banner separators used in the source file (`# ====...====`) are
    stripped before rendering. They organise the raw markdown but render as
    enormous headings in Streamlit.

    If the file is missing - which would mean the repository was assembled
    incompletely - the tab says so plainly and names the file, rather than
    rendering an empty page.
    """
    st.header("Build Notes")
    st.caption(
        "How the platform was designed and built: the transformation layer, the "
        "mart contract, the synthetic data, and what deployment taught. This is "
        "the project's build document, rendered from markdown."
    )

    doc = load_reference_doc("build_notes")
    if not doc:
        st.warning(
            f"**{BUILD_NOTES_PATH.name} was not found.**\n\n"
            "This tab renders that file from disk. It ships in the repository "
            "beside the application file; if it is missing, the deployment is "
            "incomplete."
        )
        return

    hdr_col, dl_col = st.columns([3, 1])
    with hdr_col:
        st.markdown("#### 📐 Design, build and validation")
    with dl_col:
        st.download_button(
            label="📥 Markdown",
            data=doc.encode("utf-8"),
            file_name=BUILD_NOTES_PATH.name,
            mime="text/markdown",
            key="dl_build_notes_md",
            type="primary",
        )

    st.markdown("---")
    st.markdown(_clean_reference_markdown(doc))


def _clean_reference_markdown(doc: str) -> str:
    """Strip the source file's banner separators before rendering.

    `# ====================SECTION_NAME====================` organises the raw
    file and is how these documents are written across the platform, but
    Streamlit renders it as an h1 of equals signs. The banner line is dropped;
    the human-readable heading beneath it is what the reader sees.
    """
    kept = []
    for line in doc.splitlines():
        stripped = line.strip()
        if stripped.startswith("# ====") and stripped.endswith("===="):
            continue
        kept.append(line)
    return "\n".join(kept)


# ====================SIDEBAR====================
def render_sidebar():
    """SIDEBAR, top to bottom.

        1  title      🏘️ NZ Public Housing
        2  caption    the four source agencies
        3  h4         Latest published periods
           bullets    KO stock · Housing Register · Market rent · HUD (ceased)
        4  rule
        5  h4         Real vs synthetic
           paragraph  which tabs are published data, which are modelled
        6  rule
        7  info box   independent build, not affiliated or endorsed

    The published-period list comes first because these four series end on four
    different dates. A chart that stops in 2023 is otherwise easily misread as a
    collapse rather than as a publication ending.
    """
    st.sidebar.title("🏘️ NZ Public Housing")
    st.sidebar.caption("Kāinga Ora, MSD, HUD and MBIE public data")

    snap = get_latest_snapshot(df_db_schema)
    if not snap.empty:
        row = snap.iloc[0]
        st.sidebar.markdown("#### Latest published periods")
        st.sidebar.markdown(
            f"- **Kāinga Ora stock** — {row['KO_PERIOD']}\n"
            f"- **Housing Register** — {row['REGISTER_PERIOD']}\n"
            f"- **Market rent** — {row['RENT_PERIOD']}\n"
            f"- **HUD quarterly report** — {row['CHP_PERIOD']} *(ceased)*"
        )

    st.sidebar.markdown("---")
    st.sidebar.markdown("#### Real vs synthetic")
    st.sidebar.markdown(
        "Tabs 1, 3 and 4 are built on **published data**. Property locations, "
        "tenancy rents and maintenance are **synthetic** — Kāinga Ora publishes "
        "nothing below region-by-bedroom. Synthetic tabs are banner-marked."
    )
    st.sidebar.markdown("---")
    st.sidebar.info(
        "Built by Celnic Consulting from open government data. "
        "Not affiliated with, or endorsed by, Kāinga Ora."
    )


# ====================TABS====================
# One function per tab, each drawing its block top to bottom in the order its
# docstring lists. Tabs never share mutable state; switching tabs re-runs only
# the active body against the cached queries.
def render_main_tabs():
    """The six-tab bar and its bodies, left to right.

        📊 National Overview    real, one 🔶 metric
        🗺️ Stock Map            amber banner - counts real, locations modelled
        💰 Market Rent vs IRR   mixed - real market rent, modelled IRR
        📋 Housing Register     amber banner - aggregates real, explorer modelled
        🔧 Asset & Maintenance  red banner - entirely modelled
        ⚙️ Pipeline             provenance, lineage, reconciliation
        🏗️ Build Notes          the Phase Two engineering write-up, as markdown

    Ordered so the reader meets published data before modelled data, and reaches
    the evidence for both in the final two tabs. Streamlit executes every tab body on
    each rerun, so a failure in any tab surfaces immediately rather than lying in
    wait until that tab is opened.
    """
    tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs([
        "📊 National Overview",
        "🗺️ Stock Map",
        "💰 Market Rent vs IRR",
        "📋 Housing Register",
        "🔧 Asset & Maintenance",
        "⚙️ Pipeline",
        "🏗️ Build Notes",
    ])
    with tab1:
        render_tab_national()
    with tab2:
        render_tab_map()
    with tab3:
        render_tab_rent()
    with tab4:
        render_tab_register()
    with tab5:
        render_tab_assets()
    with tab6:
        render_tab_methodology()
    with tab7:
        render_tab_build_notes()


# ---------------------------------------------------------------- tab 1
def render_tab_national():
    """TAB 1 - National Overview. Real data, one synthetic-derived metric.

        H2  National Overview                      + caption
        ---------------------------------------------------------------
        [ KO homes ][ CHP homes ][ Register ][ Median rent ][ IRRS 🔶 ]
        caption: what the 🔶 metric is, in full
        --- rule ---
        [ Public housing supply      ][ Register and transfer register ]
          line  KO homes      BRAND     line  Housing Register   INK
          line  CHP dotted    ACCENT    line  Transfer Register  pale
        h5  Kainga Ora portfolio composition
            stacked area, one band per portfolio, Blues scale
            caption: why portfolios are reported separately
        📋 National quarterly series               + 📥 Excel

    Only the fifth metric is modelled, so it alone carries 🔶 and a caption
    beneath the row explains it. A full-width banner here would wrongly imply
    the other four metrics are modelled too.

    The CHP line is dotted and stops mid-chart because HUD ceased publishing
    after 2023Q4. Dotting makes the ending read as a publication decision rather
    than as community housing disappearing.
    """
    st.header("National Overview")
    st.caption("Public housing supply, demand and market context, from the "
               "published record.")

    snap = get_latest_snapshot(df_db_schema)
    vac = get_vacancy_summary(df_db_schema)
    sub = get_subsidy_summary(df_db_schema)
    if snap.empty:
        st.warning("No data available.")
        return
    row, v, s = snap.iloc[0], vac.iloc[0], sub.iloc[0]

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Kāinga Ora homes", f"{int(row['STATE_RENTALS']):,}",
              help=f"State rentals, {row['KO_PERIOD']}. Published by Kāinga Ora.")
    c2.metric("CHP homes", f"{int(row['CHP_HOMES']):,}",
              help=f"Community Housing Provider homes, {row['CHP_PERIOD']}. "
                   f"Last published figure — HUD's quarterly report ceased after "
                   f"December 2023.")
    c3.metric("Housing Register", f"{int(row['REGISTER_TOTAL']):,}",
              help=f"Applicants on MSD's Housing Register, {row['REGISTER_PERIOD']}.")
    c4.metric("National median rent",
              f"${row['MEDIAN_RENT']:,.0f}/wk",
              help=f"All bonds, {row['RENT_PERIOD']}. MBIE bond data.")
    c5.metric("Weekly IRRS subsidy 🔶",
              f"${s['TOTAL_WEEKLY_SUBSIDY'] / 1e6:,.1f}M",
              help="SYNTHETIC. Modelled income-related rent subtracted from real "
                   "market rent. Not a published figure.")

    st.caption(
        f"🔶 The subsidy metric is **synthetic-derived**: about "
        f"\\${s['ANNUAL_SUBSIDY_BN']:.2f} billion "
        f"a year across {int(s['TENANCIES']):,} modelled tenancies, at a mean of "
        f"\\${s['AVG_SUBSIDY']:,.0f}/week. Market rent is real; the income-related "
        f"rent it is subtracted from is modelled. Every other metric above is "
        f"published data."
    )

    st.markdown("---")
    nat = get_national_quarter(df_db_schema)

    left, right = st.columns(2)
    with left:
        st.markdown("##### Public housing supply")
        fig = go.Figure()
        supply = nat[nat["KO_STATE_RENTALS"].notna()]
        fig.add_trace(go.Scatter(
            x=supply["PERIOD"], y=supply["KO_STATE_RENTALS"],
            name="Kāinga Ora homes", mode="lines",
            line=dict(color=BRAND, width=3)))
        chp = nat[nat["CHP_HOMES"].notna()]
        fig.add_trace(go.Scatter(
            x=chp["PERIOD"], y=chp["CHP_HOMES"],
            name="CHP homes (HUD, to 2023Q4)", mode="lines",
            line=dict(color=ACCENT, width=2, dash="dot")))
        fig.update_layout(height=380, hovermode="x unified",
                          yaxis_title="Homes", xaxis_title=None,
                          legend=dict(orientation="h", y=-0.2))
        st.plotly_chart(fig, width="stretch")

    with right:
        st.markdown("##### Housing Register and transfer register")
        reg = nat[nat["REGISTER_TOTAL"].notna()]
        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=reg["PERIOD"], y=reg["REGISTER_TOTAL"], name="Housing Register",
            mode="lines", line=dict(color=INK, width=3)))
        fig.add_trace(go.Scatter(
            x=reg["PERIOD"], y=reg["TRANSFER_TOTAL"], name="Transfer Register",
            mode="lines", line=dict(color="#7FB3D5", width=2)))
        fig.update_layout(height=380, hovermode="x unified",
                          yaxis_title="Applicants", xaxis_title=None,
                          legend=dict(orientation="h", y=-0.2))
        st.plotly_chart(fig, width="stretch")

    st.markdown("##### Kāinga Ora portfolio composition")
    pf = get_stock_portfolio(df_db_schema)
    if not pf.empty:
        fig = px.area(pf, x="PERIOD", y="PROPERTIES", color="PORTFOLIO",
                      color_discrete_sequence=px.colors.sequential.Blues_r)
        fig.update_layout(height=360, hovermode="x unified",
                          yaxis_title="Properties", xaxis_title=None)
        st.plotly_chart(fig, width="stretch")
        st.caption(
            "State rentals are the core social housing portfolio. Community "
            "Group Housing, the CHP lease portfolio and transitional housing are "
            "managed by Kāinga Ora but reported separately to avoid "
            "double-counting against Community Housing Provider figures."
        )

    render_detail_with_excel(
        nat, "National quarterly series",
        "Kāinga Ora / NZ Public Housing — National Quarterly Series",
        "national_quarterly", key="nat")


# ---------------------------------------------------------------- tab 2
METRIC_CHOICES = {
    "Property count": ("PROPERTY_COUNT", "blue", "properties"),
    "Mean condition score": ("MEAN_CONDITION", "blue", "score 1–5"),
    "Vacancy rate": ("VACANCY_RATE", "heat", "% vacant"),
    "Mean weekly subsidy": ("MEAN_SUBSIDY", "heat", "$/week"),
}


def render_tab_map():
    """TAB 2 - Stock Map. Counts are real, locations are modelled.

        H2  Stock Map
        🔶 AMBER banner: locations synthetic, counts real
        [ H3 resolution 8-12 ][ Colour by ][ Jump to: Wgtn/Akl/Chch/All ]
        [ Filter to a territorial authority                            ]
        ===============================================================
          pydeck H3HexagonLayer over CartoDB Voyager basemap
        ===============================================================
        caption: colour scale and its numeric range
        [ Hexagons ][ Properties ][ Mean condition ][ Mean subsidy ]
        📋 Hexagon detail, top 2,000               + 📥 Excel

    Amber rather than red: hexagon counts and their district totals are real
    published figures, and only each property's position inside its district is
    modelled. The banner says exactly that instead of implying the whole map is
    invented.

    The resolution slider carries a help note that at 11-12 nearly every
    property occupies its own hexagon, so the map stops showing density and
    starts showing the modelled point pattern - which would otherwise look like
    a finding.
    """
    st.header("Stock Map")
    render_synthetic_banner(
        "Property locations on this map are **synthetic**.",
        "Kāinga Ora publishes stock counts by territorial authority and bedroom "
        "count, and nothing finer — no addresses, no coordinates. Property counts "
        "per district and bedroom count here match the published table exactly; "
        "*where* each property sits inside its district is modelled. Auckland is "
        "distributed across its 21 local boards using Kāinga Ora's own published "
        "local-board table. Read the hexagons as a density model, not as addresses.",
        level="partial")

    c1, c2, c3 = st.columns([1, 1, 2])
    with c1:
        resolution = st.select_slider(
            "H3 resolution", options=[8, 9, 10, 11, 12], value=8,
            help="8 is roughly a suburb; 12 is roughly a building footprint. "
                 "At 11–12 nearly every property occupies its own hexagon, so "
                 "the map shows the modelled point pattern rather than density.")
    with c2:
        metric_label = st.selectbox("Colour by", list(METRIC_CHOICES))
    with c3:
        jump = st.radio("Jump to", list(QUICK_JUMPS), horizontal=True, index=0)

    ta_filter = st.selectbox("Filter to a territorial authority",
                             get_ta_options(df_db_schema), index=0)

    df = get_map_data(df_db_schema, resolution, ta_filter)
    if df.empty:
        st.info("No properties match the current filter.")
        return

    metric_col, palette, unit = METRIC_CHOICES[metric_label]
    render_h3_map(df, metric_col, palette, unit, jump, resolution)

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Hexagons", f"{len(df):,}")
    c2.metric("Properties", f"{int(df['PROPERTY_COUNT'].sum()):,}")
    c3.metric("Mean condition", f"{df['MEAN_CONDITION'].mean():.2f}")
    c4.metric("Mean subsidy", f"${df['MEAN_SUBSIDY'].mean():,.0f}/wk")

    render_detail_with_excel(
        df.sort_values("PROPERTY_COUNT", ascending=False).head(2000),
        f"Hexagon detail (H3 resolution {resolution})",
        f"Stock Map — H3 resolution {resolution} — {ta_filter}",
        "stock_map_hexagons", key="map")


def render_h3_map(df, color_column, palette, unit, jump, resolution):
    """The hexagon layer itself, plus the colour-scale caption beneath it.

    Two palettes, chosen by what the metric means rather than by taste:

        "blue"  DENSITY - pale to deep blue. More is simply more, and the eye
                reads depth as quantity without needing a legend.
        "heat"  PRESSURE - red through to yellow, red is worse. Used for
                vacancy rate and subsidy gap, where a high value is a problem
                rather than an achievement.

    Values are min-max normalised within the current filter so the ramp always
    spans the visible data, and the caption prints the real range so a reader
    can see what the darkest hexagon is worth. Zoom rises with resolution;
    otherwise a res-12 view opens far too wide to show anything.

    Tooltip on hover: dominant TA, property count, mean bedrooms, mean
    condition, vacancy rate, mean subsidy, most common dwelling type.
    """
    d = df.copy()
    values = pd.to_numeric(d[color_column], errors="coerce").fillna(0)
    lo, hi = values.min(), values.max()
    span = (hi - lo) or 1.0
    norm = ((values - lo) / span).clip(0, 1)

    if palette == "blue":
        # Blue gradient for stock density — more is simply more.
        d["COLOR"] = [[int(232 - 190 * t), int(245 - 150 * t),
                       int(253 - 90 * t), 205] for t in norm]
    else:
        # Red-to-yellow where the metric represents pressure.
        d["COLOR"] = [[240, int(60 + 195 * (1 - t)), int(40 + 40 * (1 - t)), 205]
                      for t in norm]

    lat, lon, zoom = QUICK_JUMPS[jump]
    layer = pdk.Layer(
        "H3HexagonLayer", d,
        get_hexagon="HEX",
        get_fill_color="COLOR",
        get_line_color=[255, 255, 255, 60],
        pickable=True, stroked=True, filled=True, extruded=False,
        line_width_min_pixels=1,
    )
    tooltip = {
        "html": (
            "<b>{DOMINANT_TA}</b><br/>"
            "Properties: {PROPERTY_COUNT}<br/>"
            "Mean bedrooms: {MEAN_BEDROOMS}<br/>"
            "Mean condition: {MEAN_CONDITION}<br/>"
            "Vacancy: {VACANCY_RATE}%<br/>"
            "Mean subsidy: ${MEAN_SUBSIDY}/wk<br/>"
            "Most common type: {DOMINANT_DWELLING}"
        ),
        "style": {"backgroundColor": INK, "color": "white"},
    }
    st.pydeck_chart(pdk.Deck(
        layers=[layer],
        initial_view_state=pdk.ViewState(latitude=lat, longitude=lon,
                                         zoom=zoom + (resolution - 8) * 0.55,
                                         pitch=0, bearing=0),
        tooltip=tooltip,
        map_style="https://basemaps.cartocdn.com/gl/voyager-gl-style/style.json",
    ))
    st.caption(f"Colour scale: {unit}. Range {lo:,.2f} to {hi:,.2f}.")


# ---------------------------------------------------------------- tab 3
def render_tab_rent():
    """TAB 3 - Market Rent vs Income-Related Rent. The centrepiece.

        H2  Market Rent vs Income-Related Rent     + caption
        v  ⚠️ Read this before using these numbers  EXPANDER, OPEN BY DEFAULT
           four bold caveats: provisional bond data, private-tenancy-only
           scope, TA x bedroom rent derived, income-related rent synthetic
        [ Quarter ][ Bedrooms ]
        horizontal bars, top 25 districts by market rent
            bar  market rent, derived           BRAND
            bar  income-related rent 🔶         amber, overlaid
            ◆    TA median rent, measured       INK diamonds
        h5  Subsidy gap by district
            scatter  x market rent, y income-related rent
                     bubble size = tenancies, colour = subsidy (OrRd)
            caption: how to read distance from the diagonal
        h5  Measured market rent over time
            multiselect districts -> line chart, fully measured
            caption: why geometric mean rather than arithmetic
        📋 Rent comparison                         + 📥 Excel

    The caveat expander opens by default and sits above every figure. This is
    the tab most likely to be quoted, and three of its four caveats change what
    the numbers mean rather than merely qualifying them.

    Bars are overlaid, not grouped: the subject is the gap between what the
    market would charge and what the tenant pays, and overlaying makes that gap
    the visible quantity. The measured TA median is a diamond rather than a bar
    so the eye separates it from the two modelled quantities beside it.
    """
    st.header("Market Rent vs Income-Related Rent")
    st.caption("The gap between what a state house would fetch privately and "
               "what its tenant pays — the Crown's income-related rent subsidy.")

    with st.expander("⚠️ Read this before using these numbers", expanded=True):
        st.markdown("""
**Bond data is provisional.** Tenancy Services migrated to a new bond
management system (Bond Hub, fully online from 29 June 2026). Recent periods are
sourced from both the legacy and new systems, may be incomplete, and are subject
to revision. There is a known discontinuity of roughly **17,550 additional
active bonds** arising from recording differences between the systems, so recent
periods are not strictly comparable with earlier ones.

**Market rent covers private tenancies only.** MBIE's market rent statistics
exclude government-owned properties by design — that is what makes them a valid
comparator for social housing rather than a circular one.

**Territorial authority × bedroom rent is derived, not measured.** MBIE publishes
rent by territorial authority *without* a bedroom split, and by bedroom count
against roughly 2,000 six-digit area codes for which it publishes no
code-to-name lookup. A real TA-by-bedroom market rent therefore does not exist in
the public files. The figures here scale each district's measured rent by the
national bedroom relativity for the same quarter, and are labelled **derived**
throughout.

**Income-related rent is synthetic.** No tenancy-level rent is published.
        """)

    periods = get_rent_periods(df_db_schema)
    c1, c2 = st.columns(2)
    period = c1.selectbox("Quarter", periods, index=0)
    bedrooms = c2.selectbox("Bedrooms", ["1", "2", "3", "4", "5+"], index=1)

    df = get_rent_comparison(df_db_schema, period, bedrooms)
    if df.empty:
        st.info("No comparison available for that combination.")
        return

    top = df.head(25).sort_values("MARKET_RENT")
    fig = go.Figure()
    fig.add_trace(go.Bar(
        y=top["TA_NAME"], x=top["MARKET_RENT"], name="Market rent (derived)",
        orientation="h", marker_color=BRAND))
    fig.add_trace(go.Bar(
        y=top["TA_NAME"], x=top["MEAN_IRR"], name="Income-related rent (synthetic)",
        orientation="h", marker_color="#F5B041"))
    fig.add_trace(go.Scatter(
        y=top["TA_NAME"], x=top["MEDIAN_RENT"], name="TA median rent (measured)",
        mode="markers", marker=dict(color=INK, size=9, symbol="diamond")))
    fig.update_layout(barmode="overlay", height=680, xaxis_title="NZD per week",
                      yaxis_title=None, hovermode="y unified",
                      legend=dict(orientation="h", y=-0.08))
    st.plotly_chart(fig, width="stretch")

    st.markdown("##### Subsidy gap by district")
    scatter = df[df["TENANCIES"].notna()]
    fig2 = px.scatter(
        scatter, x="MARKET_RENT", y="MEAN_IRR", size="TENANCIES",
        color="MEAN_SUBSIDY", hover_name="TA_NAME",
        color_continuous_scale="OrRd",
        labels={"MARKET_RENT": "Market rent, derived ($/week)",
                "MEAN_IRR": "Mean income-related rent, synthetic ($/week)",
                "MEAN_SUBSIDY": "Mean subsidy ($/wk)",
                "TENANCIES": "Tenancies"})
    fig2.update_layout(height=520)
    st.plotly_chart(fig2, width="stretch")
    st.caption("Bubble size is the modelled tenancy count for that district and "
               "bedroom count. Points far above the diagonal are districts where "
               "the Crown covers most of the market value of the home.")

    st.markdown("##### Measured market rent over time")
    default_tas = [t for t in ["Auckland", "Wellington City", "Christchurch City",
                               "Gisborne District"]
                   if t in set(df["TA_NAME"])]
    chosen = st.multiselect("Districts", sorted(df["TA_NAME"].unique()),
                            default=default_tas or list(df["TA_NAME"].head(3)))
    trend = get_rent_trend(df_db_schema, chosen)
    if not trend.empty:
        fig3 = px.line(trend, x="PERIOD", y="GEOMETRIC_MEAN_RENT", color="TA_NAME",
                       labels={"GEOMETRIC_MEAN_RENT": "Geometric mean rent ($/week)",
                               "PERIOD": "", "TA_NAME": "District"})
        fig3.update_layout(height=420, hovermode="x unified")
        st.plotly_chart(fig3, width="stretch")
        st.caption("Geometric mean rather than arithmetic mean: rents cluster on "
                   "round numbers, and the geometric mean is less distorted by "
                   "the long upper tail. This series is fully measured.")

    render_detail_with_excel(
        df, f"Rent comparison — {period}, {bedrooms} bedroom",
        f"Market Rent vs Income-Related Rent — {period} — {bedrooms} bedroom",
        f"rent_comparison_{period}_{bedrooms}", key="rent")


# ---------------------------------------------------------------- tab 4
def render_tab_register():
    """TAB 4 - Housing Register. Aggregates real, applicant explorer synthetic.

        H2  Housing Register
        🔶 AMBER banner: totals real, applicant explorer synthetic
        [ Register total ][ Priority A ][ Priority B ][ Transfer register ]
        stacked area  Priority A ACCENT over Priority B amber
        h5  Register by region            one line per region
        h5  Bedroom mismatch: who is waiting versus what exists
            grouped bars  demand % ACCENT  vs  KO stock % BRAND
            caption: quantifies the one-bedroom shortage
        h5  Applicant explorer 🔶 synthetic
            [ Region ][ Priority ]
            histogram  days to house, split by priority
            caption: median modelled time to house
        📋 Applicant explorer 🔶 synthetic          + 📥 Excel

    The bedroom-mismatch chart is the analytical payoff of the tab and both its
    series are real published data - demand from MSD, supply from Kainga Ora -
    so the caption says so explicitly. It is the one place where the structural
    problem is visible directly in official statistics rather than inferred.

    The register-total metric uses delta_color="inverse": a rising waitlist is
    bad news, and the default green-for-up would read as an improvement.
    """
    st.header("Housing Register")
    render_synthetic_banner(
        "Register totals, priority split and bedroom demand are **published MSD "
        "data**. The applicant-level explorer at the bottom is **synthetic**.",
        "MSD publishes the register in aggregate only. The explorer generates "
        "applicant records that reconcile exactly to the published totals for "
        "every territorial authority, priority and bedroom requirement — but no "
        "row corresponds to a real person.",
        level="partial")

    reg = get_register_trend(df_db_schema)
    if reg.empty:
        st.warning("No register data available.")
        return

    latest = reg.iloc[-1]
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Register total", f"{int(latest['REGISTER_TOTAL']):,}",
              delta=f"{int(latest['REGISTER_TOTAL'] - reg.iloc[-5]['REGISTER_TOTAL']):+,} "
                    f"vs a year earlier" if len(reg) > 5 else None,
              delta_color="inverse")
    c2.metric("Priority A", f"{int(latest['PRIORITY_A']):,}")
    c3.metric("Priority B", f"{int(latest['PRIORITY_B']):,}")
    c4.metric("Transfer Register",
              f"{int(latest['TRANSFER_TOTAL']):,}"
              if pd.notna(latest["TRANSFER_TOTAL"]) else "—")

    fig = go.Figure()
    fig.add_trace(go.Scatter(x=reg["PERIOD"], y=reg["PRIORITY_A"],
                             name="Priority A", stackgroup="one",
                             line=dict(color=ACCENT, width=0)))
    fig.add_trace(go.Scatter(x=reg["PERIOD"], y=reg["PRIORITY_B"],
                             name="Priority B", stackgroup="one",
                             line=dict(color="#F5B041", width=0)))
    fig.update_layout(height=380, hovermode="x unified",
                      yaxis_title="Applicants", xaxis_title=None,
                      legend=dict(orientation="h", y=-0.2))
    st.plotly_chart(fig, width="stretch")

    st.markdown("##### Register by region")
    byreg = get_register_by_region(df_db_schema)
    if not byreg.empty:
        fig2 = px.line(byreg, x="PERIOD", y="APPLICANTS", color="REGION",
                       labels={"APPLICANTS": "Applicants", "PERIOD": ""})
        fig2.update_layout(height=440, hovermode="x unified")
        st.plotly_chart(fig2, width="stretch")

    st.markdown("##### Bedroom mismatch: who is waiting versus what exists")
    mm = get_bedroom_mismatch(df_db_schema)
    if not mm.empty:
        fig3 = go.Figure()
        fig3.add_trace(go.Bar(x=mm["BEDROOMS"], y=mm["DEMAND_PCT"],
                              name="Register demand (%)", marker_color=ACCENT))
        fig3.add_trace(go.Bar(x=mm["BEDROOMS"], y=mm["SUPPLY_PCT"],
                              name="Kāinga Ora stock (%)", marker_color=BRAND))
        fig3.update_layout(barmode="group", height=400,
                           yaxis_title="Share of total (%)",
                           xaxis_title="Bedrooms",
                           legend=dict(orientation="h", y=-0.2))
        st.plotly_chart(fig3, width="stretch")
        one_bed = mm[mm["BEDROOMS"] == "1"]
        if not one_bed.empty and pd.notna(one_bed.iloc[0]["DEMAND_PCT"]):
            d, s = one_bed.iloc[0]["DEMAND_PCT"], one_bed.iloc[0]["SUPPLY_PCT"]
            st.caption(
                f"Both series are published data. One-bedroom homes are "
                f"**{d:.0f}%** of register demand but only **{s:.0f}%** of the "
                f"Kāinga Ora portfolio — the single-bedroom shortage is the "
                f"clearest structural mismatch in New Zealand public housing, "
                f"and it is visible directly in the official statistics."
            )

    st.markdown("##### Applicant explorer 🔶 synthetic")
    c1, c2 = st.columns(2)
    regions = ["All"] + sorted(r for r in byreg["REGION"].unique() if r)
    region = c1.selectbox("Region", regions)
    priority = c2.selectbox("Priority", ["All", "A", "B"])
    explorer = get_register_explorer(df_db_schema, region, priority)

    days = get_days_to_house(df_db_schema)
    if not days.empty:
        fig4 = px.histogram(days, x="DAYS_TO_HOUSE", color="PRIORITY", nbins=60,
                            labels={"DAYS_TO_HOUSE": "Days from lodgement to housing"},
                            color_discrete_sequence=[ACCENT, "#F5B041"])
        fig4.update_layout(height=360, bargap=0.02,
                           yaxis_title="Applicants (synthetic)")
        st.plotly_chart(fig4, width="stretch")
        st.caption(f"Median modelled time to house: "
                   f"**{days['DAYS_TO_HOUSE'].median():.0f} days**. The "
                   f"distribution is calibrated to the medians HUD published "
                   f"while its quarterly report still reported time-to-house.")

    render_detail_with_excel(
        explorer, "Applicant explorer 🔶 synthetic",
        f"Housing Register Applicant Explorer (SYNTHETIC) — {region} — priority {priority}",
        "register_applicants_synthetic", key="reg")


# ---------------------------------------------------------------- tab 5
def render_tab_assets():
    """TAB 5 - Asset & Maintenance. Entirely synthetic.

        H2  Asset & Maintenance
        🔴 RED banner, full width: nothing on this tab is measured
        [ Work orders ][ Total cost ][ Mean cost ][ Mean days to close ]
        [ Condition by build decade   ][ Maintenance cost by category  ]
          heatmap, Blues                 horizontal bars, OrRd
          caption: the bimodal build era
        h5  Mould and damp hotspots
            H3 res-8 hexagons, red-to-yellow, national view
        h5  Backlog ageing
            stacked bars by time-to-close band, split by urgency
        📋 Maintenance by district 🔶 synthetic     + 📥 Excel

    Red rather than amber, and full width: no New Zealand agency publishes
    property condition or maintenance at any grain, so unlike tabs 2 and 4 there
    is no real series underneath to anchor it. The banner states that no figure
    on the tab describes a real property or a real repair.
    """
    st.header("Asset & Maintenance")
    render_synthetic_banner(
        "Everything on this tab is **synthetic**.",
        "No New Zealand agency publishes property condition, maintenance work "
        "orders or repair costs at any grain. These 200,000 work orders are "
        "generated against the synthetic portfolio, with volume and cost "
        "correlated to modelled condition score and build era. They illustrate "
        "the shape of an asset-management dataset. **No figure on this tab "
        "describes a real property or a real repair.**",
        level="full")

    summary = get_maintenance_summary(df_db_schema)
    by_ta = get_maintenance_by_ta(df_db_schema)

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Work orders", f"{int(summary['WORK_ORDERS'].sum()):,}")
    c2.metric("Total cost", f"${summary['TOTAL_COST'].sum() / 1e6:,.1f}M")
    c3.metric("Mean cost", f"${summary['TOTAL_COST'].sum() / summary['WORK_ORDERS'].sum():,.0f}")
    c4.metric("Mean days to close",
              f"{(summary['MEAN_DAYS_TO_CLOSE'] * summary['WORK_ORDERS']).sum() / summary['WORK_ORDERS'].sum():.1f}")

    left, right = st.columns(2)
    with left:
        st.markdown("##### Condition score by build decade")
        cond = get_condition_by_decade(df_db_schema)
        pivot = cond.pivot_table(index="BUILD_DECADE", columns="CONDITION_SCORE",
                                 values="PROPERTIES", fill_value=0)
        fig = px.imshow(pivot.T, aspect="auto", color_continuous_scale="Blues",
                        labels=dict(x="Build decade", y="Condition score",
                                    color="Properties"))
        fig.update_layout(height=400)
        st.plotly_chart(fig, width="stretch")
        st.caption("The two bands reflect the modelled bimodal build history: "
                   "the post-war state-house programme and the 2019–2026 build "
                   "programme, with a trough between them.")

    with right:
        st.markdown("##### Maintenance cost by category")
        cat = summary.groupby("CATEGORY", as_index=False).agg(
            TOTAL_COST=("TOTAL_COST", "sum"), WORK_ORDERS=("WORK_ORDERS", "sum"))
        fig2 = px.bar(cat.sort_values("TOTAL_COST"), x="TOTAL_COST", y="CATEGORY",
                      orientation="h", color="TOTAL_COST",
                      color_continuous_scale="OrRd",
                      labels={"TOTAL_COST": "Total cost (NZD)", "CATEGORY": ""})
        fig2.update_layout(height=400, showlegend=False)
        st.plotly_chart(fig2, width="stretch")

    st.markdown("##### Mould and damp hotspots")
    hot = get_mould_hotspots(df_db_schema)
    if not hot.empty:
        values = hot["MOULD_ORDERS"]
        norm = ((values - values.min()) / ((values.max() - values.min()) or 1))
        hot = hot.assign(COLOR=[[240, int(60 + 195 * (1 - t)),
                                 int(40 + 40 * (1 - t)), 210] for t in norm])
        layer = pdk.Layer(
            "H3HexagonLayer", hot, get_hexagon="HEX", get_fill_color="COLOR",
            get_line_color=[255, 255, 255, 50], pickable=True, stroked=True,
            filled=True, line_width_min_pixels=1)
        st.pydeck_chart(pdk.Deck(
            layers=[layer],
            initial_view_state=pdk.ViewState(latitude=WELLINGTON[0],
                                             longitude=WELLINGTON[1],
                                             zoom=5.2),
            tooltip={"html": "<b>{DOMINANT_TA}</b><br/>Mould/damp orders: "
                             "{MOULD_ORDERS}<br/>Cost: ${TOTAL_COST}",
                     "style": {"backgroundColor": INK, "color": "white"}},
            map_style="https://basemaps.cartocdn.com/gl/voyager-gl-style/style.json"))

    st.markdown("##### Backlog ageing")
    backlog = get_backlog_ageing(df_db_schema)
    order = ["0-7 days", "8-30 days", "31-90 days", "91-180 days", "180+ days"]
    backlog["AGE_BAND"] = pd.Categorical(backlog["AGE_BAND"], order, ordered=True)
    fig3 = px.bar(backlog.sort_values("AGE_BAND"), x="AGE_BAND", y="WORK_ORDERS",
                  color="URGENCY", barmode="stack",
                  labels={"WORK_ORDERS": "Work orders", "AGE_BAND": "Time to close"})
    fig3.update_layout(height=380)
    st.plotly_chart(fig3, width="stretch")

    render_detail_with_excel(
        by_ta, "Maintenance by district 🔶 synthetic",
        "Asset & Maintenance by Territorial Authority (SYNTHETIC)",
        "maintenance_by_district_synthetic", key="maint")


# ---------------------------------------------------------------- tab 6
def render_tab_methodology():
    """TAB 6 - Pipeline. How to check everything on the other five tabs.

        H2  Pipeline & Methodology                 + intro paragraph
        H3  Reconciliation checks
            green success / red error, driven by the stored pass count
            paragraph: why cross-source checks are the ones that matter
            📋 Reconciliation checks                + 📥 Excel
        H3  Real versus synthetic lineage           📋 + 📥 Excel
        H3  Source register                         📋 + 📥 Excel
        H3  Every file downloaded                   📋 + 📥 Excel
            name, period, size, download date, MD5 checksum, licence, URL
        H3  Publication renames                     📋 + 📥 Excel
        H3  Period coverage                         📋 + 📥 Excel
        v  Known limitations, stated plainly        expander, closed

    The pass/fail state is computed from the stored validation results rather
    than asserted in prose, so the tab cannot claim the checks pass while they
    are failing.

    The download table is what the hazard banner points at: it carries a
    download date and an MD5 for all 404 source files.

    Limitations sit in a closed expander rather than being left out - available
    to anyone who looks, without displacing the reconciliation evidence above.
    """
    st.header("Pipeline & Methodology")
    st.markdown(
        "Every figure in this app is either downloaded from a New Zealand "
        "government publication or generated from one. This tab shows which is "
        "which, where each dataset came from, and how the numbers were checked."
    )

    st.markdown("### Reconciliation checks")
    val = get_validation(df_db_schema)
    passed = int((val["STATUS"] == "PASS").sum())
    if passed == len(val):
        st.success(f"All {len(val)} reconciliation checks pass.")
    else:
        st.error(f"{len(val) - passed} of {len(val)} checks are failing.")
    st.markdown(
        "The checks that matter most are the cross-source ones. A parser can "
        "agree with itself while being wrong — an early version of this pipeline "
        "dropped four territorial authority rows in five from the 2015–2019 "
        "Kāinga Ora PDFs, and the surviving rows still summed to their own totals "
        "perfectly. What caught it was comparing Kāinga Ora's stock against the "
        "figure **HUD** publishes for Kāinga Ora: two agencies, two document "
        "formats, no shared code path."
    )
    render_detail_with_excel(val, "Reconciliation checks",
                             "Reconciliation Checks", "reconciliation_checks",
                             key="val")

    st.markdown("### Real versus synthetic lineage")
    render_detail_with_excel(get_lineage(df_db_schema), "Table lineage",
                             "Real vs Synthetic Lineage", "lineage", key="lin")

    st.markdown("### Source register")
    st.caption("Most of these datasets are released under Creative Commons "
               "Attribution licences (CC BY 3.0 or 4.0 New Zealand). Check each "
               "publisher's terms before redistributing.")
    render_detail_with_excel(get_source_register(df_db_schema), "Sources",
                             "Source Register", "source_register", key="src")

    st.markdown("### Every file downloaded")
    manifest = get_download_manifest(df_db_schema)
    st.caption(f"{len(manifest):,} files across "
               f"{manifest['DATASET_ID'].nunique()} datasets.")
    render_detail_with_excel(manifest, "Download manifest",
                             "Download Manifest", "download_manifest", key="dl")

    st.markdown("### Publication renames")
    st.caption(
        "These agencies rename their own publications constantly. Left alone, "
        "that fragments one time series across several tables with holes in each. "
        "Every merge below is recorded, and the deliberate non-merges with it."
    )
    render_detail_with_excel(get_family_aliases(df_db_schema),
                             "Publication family aliases", "Family Aliases",
                             "family_aliases", key="fam")

    st.markdown("### Period coverage")
    render_detail_with_excel(get_coverage(df_db_schema), "Coverage by family",
                             "Period Coverage", "period_coverage", key="cov")

    with st.expander("Known limitations, stated plainly"):
        st.markdown("""
- **The HUD Public Housing Quarterly Report ceased after December 2023.** Its
  successor, the Government Housing Dashboard, is an embedded Power BI report
  with no export, so the CHP homes, transitional housing and Housing First series
  stop at 2023Q4. Kāinga Ora stock, the Housing Register and market rent all
  continue to 2025–2026.
- **Suburb-level bond data cannot be labelled.** MBIE's detailed quarterly file
  carries roughly 2,000 six-digit area codes and publishes no code-to-name
  lookup; its market-rent service returns suburb names without the matching
  identifier. Those rows are retained in the working database but excluded from
  the app, and TA-by-bedroom rent is derived rather than measured as a result.
- **The Market Rent API was not used.** It requires registering for an API key.
  The CSV files behind it are the same bond database, so nothing is lost.
- **Stats NZ boundaries were not used.** The Geographic Data Service requires an
  API key for every download, which would stop anyone from rebuilding this
  pipeline after cloning it. Territorial authority centroids are bundled in the
  repository instead. They are centroids, not boundaries.
- **MSD applies random rounding to base 3 and suppresses small counts.** Summing
  territorial authorities therefore understates the national total slightly —
  by up to 3% in the smallest early quarters. Suppressed cells are preserved as
  null, never as zero.
- **Nothing here is an official statistic.** This is an independent build from
  public data, not affiliated with or endorsed by Kāinga Ora, MSD, HUD or MBIE.
        """)


# ====================VISUALISATION====================
# Shared visual components used by more than one tab.
def render_synthetic_banner(headline, detail, level="partial"):
    """Per-tab synthetic-data banner. Amber for partial, red for full.

    Two levels, because the distinction is real and matters to a reader:

        "partial"  st.warning, amber - some series on this tab are published
                   data and some are modelled; the detail says which is which
        "full"     st.error, red - nothing on this tab is measured

    Placed directly under the tab heading, above every figure it qualifies.
    """
    if level == "full":
        st.error(f"🔶 **{headline}**\n\n{detail}")
    else:
        st.warning(f"🔶 **{headline}**\n\n{detail}")


# ====================STATIC_METHODS====================
def render_disclaimer_banner(
    *,
    produced_by="Celnic Consulting",
    purpose="showing the benefits of Flipping the Data Team",
    source_tab="⚙️ Pipeline",
    data_origin="NEW ZEALAND GOVERNMENT DATA",
    accent="#FFD100",
    ink="#111",
):
    """Hazard-striped provenance banner, shown above every tab.

    Reusable across Celnic data applications — nothing about this function is
    specific to housing, or to any one department. Everything that changes
    between applications is a keyword argument:

        produced_by   who built it, named as the independent party
        purpose       why it exists; set to None to drop the clause entirely
        source_tab    the tab carrying source links, dates and checksums
        data_origin   the jurisdiction, for the headline
        accent / ink  the two banner colours

    Deliberately no departments are named. Naming them invites a stale list the
    moment an application picks up another source, and the banner would then be
    quietly wrong about its own provenance. The Pipeline tab is the single place
    that enumerates sources, and it is generated from the download manifest
    rather than written by hand.

    The banner sits above the tabs, not inside a footer, because it has to be
    read before any figure is — a caveat below the chart has already failed.

    NOTE: use `st.html`, not `st.markdown(unsafe_allow_html=True)`. The latter
    strips nested divs and inline styles, which is the whole banner.
    """
    stripe = (f"height:14px; background:repeating-linear-gradient("
              f"45deg, {accent} 0 14px, {ink} 14px 28px);")
    purpose_clause = f" for the purpose of {purpose}," if purpose else ""

    st.html(
        f"""
        <div style="border:3px solid {ink}; border-radius:6px; overflow:hidden;
                    margin:0 0 14px 0; font-family:sans-serif;">
          <div style="{stripe}"></div>
          <div style="background:{accent}; color:{ink}; padding:12px 16px;">
            <div style="font-weight:800; font-size:15px; letter-spacing:.02em;">
              &#9888;&#65039; BUILT FROM {data_origin} &mdash;
              NOT AN OFFICIAL GOVERNMENT PRODUCT
            </div>
            <div style="font-size:13.5px; line-height:1.5; margin-top:6px;">
              Figures are reproduced from public releases by the departments
              listed in the <b>{source_tab}</b> tab. This application is produced
              independently by <b>{produced_by}</b>{purpose_clause} and
              <b>does not represent the views, policy or official statistics of
              those departments</b>.
              <b>Please refer to the original figures in the data source and do
              not rely on these.</b>
              Every original source file, with its download date and checksum,
              is listed in the <b>{source_tab}</b> tab.
            </div>
          </div>
          <div style="{stripe}"></div>
        </div>
        """
    )


def render_attribution():
    """Bottom-of-page attribution expander, drawn under every tab.

    The hazard banner above the tabs carries the warning; this carries the
    licence. Both are needed and neither replaces the other: CC BY requires the
    attribution to travel with the work, and a reader who scrolls to the bottom
    of a chart should be able to find out whose data it is without opening the
    Pipeline tab.

    Closed by default, because the banner has already said the loud part.
    """
    with st.expander("Data sources & attribution"):
        st.markdown(
            "Built on open data used under **CC BY 4.0** — modified and "
            "partly synthetic; demonstration of method, not published "
            "statistics. Full provenance: "
            "[ATTRIBUTION.md](https://github.com/celnicconsulting/kainga_ora_housing_nz/blob/main/ATTRIBUTION.md)."
        )


def _safe_filename(text):
    """Sanitise text for use in a download filename.

    Keeps alphanumerics, spaces and dashes, converts spaces to underscores and
    truncates to 40 characters. Applied to anything that reaches a filename from
    a user-facing selection, such as the quarter and bedroom count on tab 3.
    """
    cleaned = "".join(c for c in str(text) if c.isalnum() or c in " _-").strip()
    return cleaned.replace(" ", "_")[:40] or "results"


def build_styled_excel(df, title, sheet_name="Results"):
    """Generic styled in-memory Excel export for any dataframe."""
    df = df.head(50000)
    cols = list(df.columns)
    ncols = max(len(cols), 1)
    wrap_cols = {"NOTES", "NOTE", "DETAIL", "RATIONALE", "URL", "SOURCE"}

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    title_cell.fill = PatternFill("solid", start_color="1A3A5C")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    hdr_fill = PatternFill("solid", start_color="2E86AB")
    hdr_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    hdr_border = Border(bottom=Side(style="medium", color="1A3A5C"))
    for ci, col in enumerate(cols, start=1):
        cell = ws.cell(row=2, column=ci, value=str(col))
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = hdr_border
    ws.row_dimensions[2].height = 18

    is_dt = {col: pd.api.types.is_datetime64_any_dtype(df[col]) for col in cols}
    is_num = {col: pd.api.types.is_numeric_dtype(df[col]) for col in cols}
    alt_fill = PatternFill("solid", start_color="EBF5FB")
    thin_border = Border(bottom=Side(style="thin", color="D5D8DC"))

    for ri, (_, row) in enumerate(df.iterrows(), start=3):
        row_fill = alt_fill if ri % 2 == 0 else None
        for ci, col in enumerate(cols, start=1):
            val = row[col]
            cell = ws.cell(row=ri, column=ci)
            if (pd.isna(val) if not isinstance(val, str) else val == ""):
                cell.value = None
            elif is_dt[col]:
                try:
                    cell.value = pd.to_datetime(val).to_pydatetime()
                    cell.number_format = "DD-MMM-YY"
                except Exception:
                    cell.value = str(val)
            elif is_num[col]:
                try:
                    fv = float(val)
                    cell.value = int(fv) if fv.is_integer() else round(fv, 4)
                except Exception:
                    cell.value = str(val)
            else:
                cell.value = str(val)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=(col in wrap_cols))
            cell.border = thin_border
            if row_fill:
                cell.fill = row_fill

    for ci, col in enumerate(cols, start=1):
        try:
            sample = df[col].head(200).astype(str).map(len).max()
            maxlen = max(len(str(col)), int(sample) if pd.notna(sample) else 0)
        except Exception:
            maxlen = len(str(col))
        ws.column_dimensions[get_column_letter(ci)].width = min(max(maxlen + 2, 10), 45)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(ncols)}2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def render_detail_with_excel(df, heading, excel_title, file_stem, key):
    """Detail dataframe with the styled Excel button on the title row."""
    hdr_col, dl_col = st.columns([3, 1])
    with hdr_col:
        st.markdown(f"#### 📋 {heading}")
    with dl_col:
        if not df.empty:
            st.download_button(
                label="📥 Excel",
                data=build_styled_excel(df, excel_title, heading[:31]),
                file_name=f"{_safe_filename(file_stem)}.xlsx",
                mime="application/vnd.openxmlformats-officedocument."
                     "spreadsheetml.sheet",
                key=f"dl_{key}_xlsx",
                type="primary",
            )
    if df.empty:
        st.info("No data available.")
    else:
        st.dataframe(df, width="stretch", hide_index=True)


# ====================MAIN====================
def main():
    """Draw the page, top to bottom.

        1  sidebar               published periods, real-vs-synthetic note
        2  H1                    application title
        3  hazard banner         provenance warning, above the tabs
        4  tab bar and body      the six tabs
        5  attribution           CC BY licence expander, below the tabs

    The hazard banner is drawn before the tabs, not inside a footer, so the
    provenance caveat is read before any figure rather than found underneath
    one. The attribution expander is the licence, not the warning, and belongs
    at the foot of the page where a citation would go.
    """
    render_sidebar()
    st.title("Kāinga Ora / New Zealand Public Housing Intelligence")
    render_disclaimer_banner()
    render_main_tabs()
    render_attribution()


if __name__ == "__main__":
    main()
